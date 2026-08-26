from openpyxl import load_workbook


class ExcelSource:
    def __init__(self, filename):
        self.filename = filename

    def read_table(self, table_name):
        wb = load_workbook(
            self.filename,
            data_only=False
        )

        for ws in wb.worksheets:

            if table_name not in ws.tables:
                continue

            table = ws.tables[table_name]

            rows = ws[table.ref]

            headers = [
                cell.value
                for cell in rows[0]
            ]

            records = [
                tuple(cell.value for cell in row)
                for row in rows[1:]
            ]

            return headers, records

        raise ValueError(
            f"Table not found: {table_name}"
        )

    def read_sheet(self, sheet_name=None, header_row=1):
        """
        Read data from any sheet by specifying the header row index (1-based).
        """
        wb = load_workbook(
            self.filename,
            data_only=True
        )

        if sheet_name is None:
            ws = wb.active
        else:
            ws = wb[sheet_name]

        all_rows = list(ws.iter_rows(values_only=True))

        if not all_rows or header_row < 1 or header_row > len(all_rows):
            raise ValueError(
                f"header_row {header_row} is out of bounds."
            )

        header_cells = all_rows[header_row - 1]

        # Stop headers at last non-empty column
        last_col = 0
        for idx, cell in enumerate(header_cells):
            if cell is not None and str(cell).strip() != "":
                last_col = idx + 1

        headers = [
            str(cell).strip() if cell is not None else f"Column_{idx+1}"
            for idx, cell in enumerate(header_cells[:last_col])
        ]

        records = []
        for row in all_rows[header_row:]:
            row_data = row[:last_col]
            if any(cell is not None for cell in row_data):
                records.append(tuple(row_data))

        return headers, records